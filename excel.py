from openpyxl import load_workbook
import openpyxl

def excel():
    wb = load_workbook('D:/Users/Jrodriguezc/OneDrive - Linea Directa S.A.S/Escritorio/Scripting_2/Libro1.xlsx')
    sheet = wb ["Hoja3"]
    

    for i in range (3,sheet.max_row):
        if sheet.cell(row=i,column=2).value==None:
            return i 
        print(i)





