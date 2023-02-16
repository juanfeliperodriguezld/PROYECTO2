import win32com.client as win32
from openpyxl import load_workbook
from log_clase import log 
from datetime import datetime
import sys
import time

 
def conexion ():
    
    SapGuiAuto  = win32.GetObject("SAPGUI")
    if not type(SapGuiAuto) == win32.CDispatch:
        return

    application = SapGuiAuto.GetScriptingEngine
    if not type(application) == win32.CDispatch:
        SapGuiAuto = None
        return
    
    connection = application.Children(0)
    if not type(connection) == win32.CDispatch:
        application = None
        SapGuiAuto = None
        return
        
    session = connection.Children(0)
    if not type(session) == win32.CDispatch:
        connection = None
        application = None
        SapGuiAuto = None
        return
    return session


# lectura hoja de excel y funcion principal 
def registro_material(ruta_archivo):

    try:
        wb = load_workbook(ruta_archivo)

        sheet = wb["Hoja3"]  # buscando dentro del archivo la hoja

        for i in range(3, sheet.max_row + 1):  # recorre la hoja
            # valida que el valor de la fila sea un valor valido
            if sheet.cell(row=i, column=2).value == None:
                break



            rn = datetime.now()
            logs.idtransaccion = datetime.now().strftime("%H:%M:%S-")+str(rn.microsecond)
            logs.fechainicio = datetime.now()
            logs.level = 'INFO'
            logs.procesointerno = 'Inicio Transaccion'
            logs.mensaje = 'Inicio Transaccion'
            #logs.post()

            session.findById("wnd[0]").maximize()
            session.findById("wnd[0]/tbar[0]/okcd").text = "/NMEK32"

            logs.level = 'INFO'
            logs.procesointerno = 'NMEK32'
            logs.mensaje = 'Ingresando Valores'
            #logs.post()


            session.findById("wnd[0]").sendVKey (0)
            session.findById("wnd[0]/usr/ctxtRV13A-KSCHL").text = "j3ap"
            session.findById("wnd[0]/usr/ctxtRV13A-KSCHL").caretPosition = 4
            session.findById("wnd[0]").sendVKey (0)
            session.findById("wnd[0]/tbar[1]/btn[17]").press()
            session.findById("wnd[1]/usr/sub:SAPLV14A:0100/radRV130-SELKZ[8,0]").select()
            session.findById("wnd[1]/usr/sub:SAPLV14A:0100/radRV130-SELKZ[8,0]").setFocus()
            session.findById("wnd[1]/tbar[0]/btn[0]").press()
            session.findById("wnd[0]/usr/ctxtF001").text = sheet.cell(row=i, column=2).value
            session.findById("wnd[0]/usr/ctxtF002").text = sheet.cell(row=i, column=3).value
            session.findById("wnd[0]/usr/ctxtF003").text = sheet.cell(row=i, column=1).value
            session.findById("wnd[0]/usr/ctxtF004").text = "0"
            session.findById("wnd[0]/usr/ctxtF005-LOW").text = sheet.cell(row=i, column=4).value
            session.findById("wnd[0]/usr/ctxtF005-LOW").setFocus()
            session.findById("wnd[0]/usr/ctxtF005-LOW").caretPosition = 4
            session.findById("wnd[0]").sendVKey (0)
            session.findById("wnd[0]/tbar[1]/btn[8]").press()
            session.findById("wnd[0]/mbar/menu[0]/menu[1]").select()
            session.findById("wnd[0]/usr/ctxtF001").text = sheet.cell(row=i, column=2).value
            session.findById("wnd[0]/usr/ctxtF002").text = sheet.cell(row=i, column=3).value
            session.findById("wnd[0]/usr/ctxtF003").text = sheet.cell(row=i, column=1).value
            session.findById("wnd[0]/usr/ctxtF004").text = "0"
            session.findById("wnd[0]/usr/ctxtF005-LOW").text = sheet.cell(row=i, column=4).value
            session.findById("wnd[0]/usr/ctxtF005-LOW").setFocus()
            session.findById("wnd[0]/usr/ctxtF005-LOW").caretPosition = 4
            session.findById("wnd[0]").sendVKey (0)
            session.findById("wnd[0]/tbar[1]/btn[8]").press()
            session.findById("wnd[0]/usr/tblSAPMV13ATCTRL_FAST_ENTRY/ctxtKOMG-J_3ASIZE[0,0]").text = sheet.cell(row=i, column=4).value
            session.findById("wnd[0]/usr/tblSAPMV13ATCTRL_FAST_ENTRY/txtKONP-KBETR[2,0]").text = sheet.cell(row=i, column=10).value
            session.findById("wnd[0]/usr/tblSAPMV13ATCTRL_FAST_ENTRY/ctxtRV13A-DATAB[8,0]").text = sheet.cell(row=i, column=7).value
            session.findById("wnd[0]/usr/tblSAPMV13ATCTRL_FAST_ENTRY/ctxtRV13A-DATBI[9,0]").text = sheet.cell(row=i, column=8).value
            session.findById("wnd[0]/usr/tblSAPMV13ATCTRL_FAST_ENTRY/ctxtRV13A-DATBI[9,0]").setFocus()
            session.findById("wnd[0]/usr/tblSAPMV13ATCTRL_FAST_ENTRY/ctxtRV13A-DATBI[9,0]").caretPosition = 0
            session.findById("wnd[0]").sendVKey (0)
            session.findById("wnd[0]/tbar[0]/btn[0]").press()
            session.findById("wnd[0]/usr/tblSAPMV13ATCTRL_FAST_ENTRY").getAbsoluteRow(0).selected = True
            session.findById("wnd[0]/usr/tblSAPMV13ATCTRL_FAST_ENTRY/ctxtKOMG-J_3ASIZE[0,0]").setFocus()
            session.findById("wnd[0]/usr/tblSAPMV13ATCTRL_FAST_ENTRY/ctxtKOMG-J_3ASIZE[0,0]").caretPosition = 0
            session.findById("wnd[0]/tbar[1]/btn[2]").press()
            session.findById("wnd[0]/usr/tblSAPMV13ATCTRL_D0303/txtKONM-KSTBM[1,0]").text = sheet.cell(row=i, column=9).value
            session.findById("wnd[0]/usr/tblSAPMV13ATCTRL_D0303/txtKONM-KSTBM[1,1]").text = sheet.cell(row=i, column=11).value
            session.findById("wnd[0]/usr/tblSAPMV13ATCTRL_D0303/txtKONM-KSTBM[1,2]").text = sheet.cell(row=i, column=13).value
            session.findById("wnd[0]/usr/tblSAPMV13ATCTRL_D0303/txtKONM-KBETR[3,0]").text = sheet.cell(row=i, column=10).value
            session.findById("wnd[0]/usr/tblSAPMV13ATCTRL_D0303/txtKONM-KBETR[3,1]").text = sheet.cell(row=i, column=12).value
            session.findById("wnd[0]/usr/tblSAPMV13ATCTRL_D0303/txtKONM-KBETR[3,2]").text = sheet.cell(row=i, column=14).value
            session.findById("wnd[0]/usr/tblSAPMV13ATCTRL_D0303/txtKONM-KBETR[3,2]").setFocus()
            session.findById("wnd[0]/usr/tblSAPMV13ATCTRL_D0303/txtKONM-KBETR[3,2]").caretPosition = 12
            session.findById("wnd[0]").sendVKey (0)
            session.findById("wnd[0]/tbar[0]/btn[0]").press()
            #session.findById("wnd[0]/tbar[0]/btn[11]").press()
            #session.findById("wnd[1]").close
            session.findById("wnd[0]/tbar[0]/btn[11]").press()
            session.findById("wnd[1]/tbar[0]/btn[5]").press()
       
            # da respuesta a materiales modificados en el archivo excel
            sheet.cell(row=i, column=15).value = session.findById(
                "wnd[0]/sbar").text
            wb.save(ruta_archivo)  # guardar archivo excel
            time.sleep(3)

            logs.level = 'INFO'
            logs.procesointerno = 'Fin Transaccion'
            logs.mensaje = 'Fin Transaccion'
            logs.fintransaccion = datetime.now()
            #logs.postFin()

            print(session.findById("wnd[0]/sbar").text)


    except Exception as e:
        print(e)


# adminstra y llama a las funciones
def main(ruta_archivo):
    global logs
    logs = log()
    global session
    session = conexion()
    registro_material(ruta_archivo)